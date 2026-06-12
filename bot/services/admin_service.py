from __future__ import annotations

import os
from datetime import date, datetime, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from database.db import get_connection


# ─── ЗАПИСИ ──────────────────────────────────────────────────────────────────

def get_bookings_by_date(d: date) -> list:
    with get_connection() as conn:
        return conn.execute(
            """
            SELECT b.id, b.date, b.start_time, b.end_time, b.status,
                   u.first_name, u.last_name, u.phone, u.username
            FROM bookings b
            JOIN users u ON u.user_id = b.user_id
            WHERE b.date = ? AND b.status IN ('active','confirmed')
            ORDER BY b.start_time
            """,
            (d.isoformat(),),
        ).fetchall()


def get_future_bookings() -> list:
    with get_connection() as conn:
        return conn.execute(
            """
            SELECT b.id, b.date, b.start_time, b.end_time, b.status,
                   u.first_name, u.last_name, u.phone, u.username, u.user_id
            FROM bookings b
            JOIN users u ON u.user_id = b.user_id
            WHERE b.date >= ? AND b.status IN ('active','confirmed')
            ORDER BY b.date, b.start_time
            """,
            (date.today().isoformat(),),
        ).fetchall()


def get_booking_by_id(booking_id: int):
    with get_connection() as conn:
        return conn.execute(
            """
            SELECT b.id, b.date, b.start_time, b.end_time, b.status,
                   u.first_name, u.last_name, u.phone, u.username, u.user_id
            FROM bookings b
            JOIN users u ON u.user_id = b.user_id
            WHERE b.id = ?
            """,
            (booking_id,),
        ).fetchone()


def get_bookings_for_export(period: str) -> list:
    today = date.today()
    if period == "week":
        end = today + timedelta(days=6 - today.weekday())
    else:  # month
        if today.month == 12:
            end = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            end = date(today.year, today.month + 1, 1) - timedelta(days=1)

    with get_connection() as conn:
        return conn.execute(
            """
            SELECT b.date, b.start_time, b.end_time,
                   u.first_name, u.last_name, u.phone, u.username, b.status
            FROM bookings b
            JOIN users u ON u.user_id = b.user_id
            WHERE b.date BETWEEN ? AND ?
              AND b.status IN ('active','confirmed')
            ORDER BY b.date, b.start_time
            """,
            (today.isoformat(), end.isoformat()),
        ).fetchall()


# ─── РЕДАКТИРОВАНИЕ ЗАПИСИ ────────────────────────────────────────────────────

BREAK_DURATION = timedelta(minutes=30)
SESSION_DURATION = timedelta(minutes=60)


def _has_conflict(booking_id: int, d: date, new_start: datetime, new_end: datetime) -> bool:
    """Проверяет конфликт с другими записями (исключая саму себя)."""
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT start_time, end_time FROM bookings
            WHERE date = ? AND id != ? AND status IN ('active','confirmed')
            """,
            (d.isoformat(), booking_id),
        ).fetchall()

    for row in rows:
        b_start = datetime.combine(d, _t(row["start_time"]))
        b_end   = datetime.combine(d, _t(row["end_time"])) + BREAK_DURATION
        if new_start < b_end and new_end > b_start:
            return True
    return False


def _t(s: str):
    from datetime import time
    return time.fromisoformat(s)


def shift_booking(booking_id: int, delta_minutes: int) -> tuple[bool, str]:
    """
    Сдвигает start_time и end_time на delta_minutes.
    Возвращает (success, error_message).
    """
    booking = get_booking_by_id(booking_id)
    if not booking:
        return False, "Запись не найдена."

    d = date.fromisoformat(booking["date"])
    old_start = datetime.combine(d, _t(booking["start_time"]))
    old_end   = datetime.combine(d, _t(booking["end_time"]))
    delta     = timedelta(minutes=delta_minutes)

    new_start = old_start + delta
    new_end   = old_end   + delta

    # Проверяем рабочие часы
    from database.logic import get_working_hours
    hours = get_working_hours(d)
    if hours is None:
        return False, "День не рабочий."
    work_start, work_end = hours
    if new_start.time() < work_start or new_end.time() > work_end:
        return False, "Время выходит за пределы рабочего дня."

    if _has_conflict(booking_id, d, new_start, new_end):
        return False, "Конфликт с другой записью."

    with get_connection() as conn:
        conn.execute(
            "UPDATE bookings SET start_time = ?, end_time = ? WHERE id = ?",
            (new_start.time().isoformat(), new_end.time().isoformat(), booking_id),
        )
    return True, ""


def extend_booking(booking_id: int, delta_minutes: int) -> tuple[bool, str]:
    """
    Продлевает или сокращает end_time на delta_minutes.
    """
    booking = get_booking_by_id(booking_id)
    if not booking:
        return False, "Запись не найдена."

    d         = date.fromisoformat(booking["date"])
    start_dt  = datetime.combine(d, _t(booking["start_time"]))
    old_end   = datetime.combine(d, _t(booking["end_time"]))
    new_end   = old_end + timedelta(minutes=delta_minutes)

    if new_end <= start_dt + timedelta(minutes=30):
        return False, "Минимальная длительность — 30 минут."

    from database.logic import get_working_hours
    hours = get_working_hours(d)
    if hours and new_end.time() > hours[1]:
        return False, "Время выходит за пределы рабочего дня."

    if _has_conflict(booking_id, d, start_dt, new_end):
        return False, "Конфликт с другой записью."

    with get_connection() as conn:
        conn.execute(
            "UPDATE bookings SET end_time = ? WHERE id = ?",
            (new_end.time().isoformat(), booking_id),
        )
    return True, ""


# ─── РАСПИСАНИЕ ───────────────────────────────────────────────────────────────

def get_full_template() -> dict:
    """Возвращает {weekday: Row} для всех 7 дней."""
    with get_connection() as conn:
        rows = conn.execute(
            "SELECT * FROM work_schedule_template ORDER BY weekday"
        ).fetchall()
    return {row["weekday"]: row for row in rows}


def update_schedule_day(
    weekday: int,
    is_working: bool,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
) -> None:
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO work_schedule_template (weekday, start_time, end_time, is_working_day)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(weekday) DO UPDATE SET
                start_time     = excluded.start_time,
                end_time       = excluded.end_time,
                is_working_day = excluded.is_working_day
            """,
            (weekday, start_time or "10:00", end_time or "18:00", int(is_working)),
        )


def get_exceptions(from_date: Optional[date] = None) -> list:
    d = (from_date or date.today()).isoformat()
    with get_connection() as conn:
        return conn.execute(
            "SELECT * FROM work_exceptions WHERE date >= ? ORDER BY date",
            (d,),
        ).fetchall()


def add_exception(
    d: date,
    is_day_off: bool,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
) -> None:
    with get_connection() as conn:
        conn.execute(
            """
            INSERT INTO work_exceptions (date, is_day_off, start_time, end_time)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(date) DO UPDATE SET
                is_day_off = excluded.is_day_off,
                start_time = excluded.start_time,
                end_time   = excluded.end_time
            """,
            (d.isoformat(), int(is_day_off), start_time, end_time),
        )


def delete_exception(d: date) -> None:
    with get_connection() as conn:
        conn.execute("DELETE FROM work_exceptions WHERE date = ?", (d.isoformat(),))


def add_vacation(start: date, end: date) -> None:
    """Создаёт is_day_off=1 для каждого дня диапазона."""
    current = start
    while current <= end:
        add_exception(current, is_day_off=True)
        current += timedelta(days=1)


# ─── ЭКСПОРТ EXCEL ────────────────────────────────────────────────────────────

def export_bookings_to_excel(period: str) -> str:
    rows = get_bookings_for_export(period)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Записи"

    headers = ["Дата", "Начало", "Конец", "Имя", "Фамилия", "Телефон", "Username", "Статус"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([
            row["date"],
            row["start_time"][:5],
            row["end_time"][:5],
            row["first_name"],
            row["last_name"],
            row["phone"] or "—",
            f"@{row['username']}" if row["username"] else "—",
            row["status"],
        ])

    # Автоширина колонок
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_len + 4

    filename = f"bookings_export_{date.today().isoformat()}.xlsx"
    path = os.path.join(os.path.dirname(__file__), "..", "..", filename)
    path = os.path.normpath(path)
    wb.save(path)
    return path